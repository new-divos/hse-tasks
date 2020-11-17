#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from collections import namedtuple
from pathlib import Path

import openpyxl

if __name__ == '__main__':
    Person = namedtuple('Person', ('name', 'salary'))

    persons = []

    root = Path(__file__).absolute().parent.parent.parent.parent
    root = root / 'data' / 'rogaikopyta'
    for path in root.iterdir():
        if path.is_file():
            wb = openpyxl.load_workbook(path)

            sheet = wb.active
            person = Person(
                name=sheet.cell(row=2, column=2).value,
                salary=sheet.cell(row=2, column=4).value,
            )
            persons.append(person)

    assert len(persons) == 1000

    persons.sort(key=lambda item: item.name)
    for person in persons:
        print(f"{person.name} {person.salary}")
